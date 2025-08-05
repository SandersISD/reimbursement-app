#!/usr/bin/env python3

import os
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import app
from utils import generate_excel_isd_report
from models import db, Claim
import openpyxl

def test_new_template():
    with app.app_context():
        # Get claims from May 2025
        claims = Claim.query.filter(
            db.extract('month', Claim.from_date) == 5,
            db.extract('year', Claim.from_date) == 2025
        ).all()

        print(f'Found {len(claims)} claims')
        for claim in claims:
            from models import ClaimItem
            items = ClaimItem.query.filter_by(claim_id=claim.claim_id).all()
            print(f'Claim {claim.claim_id}: {claim.from_date} - {len(items)} items')

        # Generate Excel report
        excel_data = generate_excel_isd_report('2025-05')
        if excel_data:
            with open('test_new_template.xlsx', 'wb') as f:
                f.write(excel_data)
            print('✅ Excel file generated: test_new_template.xlsx')
            
            # Now analyze the generated file
            wb = openpyxl.load_workbook('test_new_template.xlsx')
            ws = wb.active
            
            print(f'Generated file has {ws.max_row} rows')
            print('\nContent around data area:')
            for row in range(10, min(35, ws.max_row + 1)):
                values = []
                for col in range(1, 8):  # A to G columns
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        values.append(f'{openpyxl.utils.get_column_letter(col)}: {str(cell.value)[:30]}')
                if values:
                    print(f'Row {row:2d}: {" | ".join(values)}')
                elif row >= 10 and row <= 30:
                    print(f'Row {row:2d}: [EMPTY]')
        else:
            print('❌ Failed to generate Excel report')

if __name__ == '__main__':
    test_new_template()
