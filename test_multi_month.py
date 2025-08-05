#!/usr/bin/env python3

import os
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import app
from utils import generate_excel_isd_report
from models import db, Claim, ClaimItem
import openpyxl

def test_multi_month_template():
    with app.app_context():
        # Get claims from both May and June 2025
        claims = Claim.query.filter(
            db.or_(
                db.and_(
                    db.extract('month', Claim.from_date) == 5,
                    db.extract('year', Claim.from_date) == 2025
                ),
                db.and_(
                    db.extract('month', Claim.from_date) == 6,
                    db.extract('year', Claim.from_date) == 2025
                )
            )
        ).all()

        print(f'Found {len(claims)} claims across multiple months')
        for claim in claims:
            items = ClaimItem.query.filter_by(claim_id=claim.claim_id).all()
            print(f'Claim {claim.claim_id}: {claim.from_date} - {len(items)} items')

        # Generate Excel report for multiple months
        excel_data = generate_excel_isd_report(claims)
        if excel_data:
            with open('test_multi_month_template.xlsx', 'wb') as f:
                f.write(excel_data)
            print('✅ Excel file generated: test_multi_month_template.xlsx')
            
            # Now analyze the generated file
            wb = openpyxl.load_workbook('test_multi_month_template.xlsx')
            ws = wb.active
            
            print(f'Generated file has {ws.max_row} rows')
            print('\nContent structure:')
            for row in range(1, min(50, ws.max_row + 1)):
                values = []
                for col in range(1, 8):  # A to G columns
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        values.append(f'{openpyxl.utils.get_column_letter(col)}: {str(cell.value)[:30]}')
                if values:
                    print(f'Row {row:2d}: {" | ".join(values)}')
                elif row >= 10 and row <= 40 and row % 5 == 0:  # Show some empty rows for structure
                    print(f'Row {row:2d}: [EMPTY]')
        else:
            print('❌ Failed to generate Excel report')

if __name__ == '__main__':
    test_multi_month_template()
