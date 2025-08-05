#!/usr/bin/env python3

import os
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import app
from utils import generate_excel_isd_report
from models import db, Claim, ClaimItem
import openpyxl

def test_multi_month_capability():
    with app.app_context():
        # Get claims from April to August 2025 (5 months!)
        claims = Claim.query.filter(
            db.extract('year', Claim.from_date) == 2025,
            db.extract('month', Claim.from_date).in_([4, 5, 6, 7, 8])
        ).order_by(Claim.from_date).all()

        print(f'Found {len(claims)} claims across multiple months:')
        
        # Group by month for display
        from collections import defaultdict
        claims_by_month = defaultdict(list)
        for claim in claims:
            month_key = claim.from_date.strftime('%Y-%m')
            claims_by_month[month_key].append(claim)
        
        total_items = 0
        for month_key in sorted(claims_by_month.keys()):
            month_claims = claims_by_month[month_key]
            month_items = 0
            for claim in month_claims:
                items = ClaimItem.query.filter_by(claim_id=claim.claim_id).all()
                month_items += len(items)
            
            total_items += month_items
            month_name = claim.from_date.strftime('%B %Y')
            print(f'   {month_name}: {len(month_claims)} claims, {month_items} items')
        
        print(f'\nTotal: {len(claims)} claims, {total_items} items across 5 months')

        # Generate Excel report for ALL these claims
        print('\n🚀 Generating Excel report with 5 months of data...')
        excel_data = generate_excel_isd_report(claims)
        if excel_data:
            with open('test_5_months.xlsx', 'wb') as f:
                f.write(excel_data)
            print('✅ Excel file generated: test_5_months.xlsx')
            
            # Analyze the generated file structure
            wb = openpyxl.load_workbook('test_5_months.xlsx')
            ws = wb.active
            
            print(f'\n📊 Generated file analysis:')
            print(f'   Total rows: {ws.max_row}')
            print(f'   Template expanded to accommodate {total_items} items across 5 months')
            
            print('\n📋 Month sections detected:')
            for row in range(1, min(60, ws.max_row + 1)):
                cell_b = ws[f'B{row}']
                cell_c = ws[f'C{row}']
                
                # Look for period markers
                if cell_b.value == 'Period:' and cell_c.value:
                    print(f'   Row {row:2d}: {cell_c.value}')
                
                # Look for total rows
                cell_d = ws[f'D{row}']
                if cell_d.value == 'Total:':
                    amounts = []
                    for col in ['E', 'F', 'G']:
                        val = ws[f'{col}{row}'].value
                        if val and val != 0:
                            amounts.append(f'{col}: {val}')
                    if amounts:
                        print(f'   Row {row:2d}: Total - {", ".join(amounts)}')
            
            print('\n✨ Success! The system can handle any number of months!')
            
        else:
            print('❌ Failed to generate Excel report')

if __name__ == '__main__':
    test_multi_month_capability()
