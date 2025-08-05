#!/usr/bin/env python3
"""
Analyze the Excel template structure
"""
import openpyxl
import os

def analyze_template():
    template_path = 'reference/!ISD Reimbursement Form.xlsx'
    
    if not os.path.exists(template_path):
        print(f"❌ Template not found: {template_path}")
        return
    
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        print('📊 ISD REIMBURSEMENT FORM TEMPLATE STRUCTURE')
        print('=' * 55)
        print(f'Sheet Name: "{ws.title}"')
        print(f'Dimensions: {ws.max_row} rows × {ws.max_column} columns')
        print()
        
        # Form header analysis
        print('🏷️  FORM HEADER SECTION:')
        header_info = [
            ('B2', 'Form Title'),
            ('B3', 'E-Form Label'),
            ('C3', 'SE Number'),
            ('B4', 'Claimant Label'),
            ('C4', 'Claimant Name'),
            ('B5', 'ID Label'), 
            ('C5', 'Student/Staff ID'),
            ('B6', 'Email Label'),
            ('C6', 'Email Address'),
            ('E4', 'Project No Label'),
            ('E5', 'Supervisor Label'),
            ('B10', 'Period Label'),
            ('C10', 'Period Value')
        ]
        
        for cell_ref, description in header_info:
            cell_value = ws[cell_ref].value
            if cell_value:
                print(f'  {cell_ref:3} ({description:15}): {cell_value}')
        
        print()
        print('📋 DATA TABLE STRUCTURE:')
        print('Headers (Row 11):')
        
        # Table column analysis
        table_columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        for col in table_columns:
            cell_ref = f'{col}11'
            header = ws[cell_ref].value
            if header:
                print(f'  Column {col}: {header}')
        
        print()
        print('📝 SAMPLE DATA ROWS:')
        # Show a few sample data rows
        for row in range(12, min(20, ws.max_row + 1)):
            row_data = []
            has_data = False
            for col in table_columns:
                cell_value = ws[f'{col}{row}'].value
                if cell_value is not None:
                    has_data = True
                    row_data.append(f'{col}:{cell_value}')
            
            if has_data:
                print(f'  Row {row:2}: {" | ".join(row_data)}')
        
        print()
        print('💡 TEMPLATE MAPPING FOR AUTOMATION:')
        print('  Personal Info Section (Rows 2-10):')
        print('    - C3: SE/E-Form Number')
        print('    - C4: Claimant Name')  
        print('    - C5: Student/Staff ID')
        print('    - C6: Email Address')
        print('    - E4: Project Number')
        print('    - E5: Supervisor Name')
        print('    - C10: Period (Date range)')
        print()
        print('  Data Table (Starting Row 12):')
        print('    - Column B: Receipt Order')
        print('    - Column C: Payment Date')
        print('    - Column D: Particulars')
        print('    - Column E: HKD Amount')
        print('    - Column F: RMB Amount')
        print('    - Column G: Other Currency (e.g., EUR)')
        print('    - Column H: Receipt Attached')
        
        print()
        print('✅ Template analysis complete!')
        print('   Ready for data export implementation!')
        
    except Exception as e:
        print(f'❌ Error analyzing template: {e}')
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    analyze_template()
