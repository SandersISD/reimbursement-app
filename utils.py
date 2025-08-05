"""
Utility functions for the Reimbursement Application
"""
import os
import csv
import zipfile
from io import StringIO
from werkzeug.utils import secure_filename
from datetime import datetime


def allowed_file(filename):
    """Check if file extension is allowed"""
    ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg'}
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def save_uploaded_file(file, upload_folder, claim_id):
    """Save uploaded file with secure filename and return the path"""
    if file and allowed_file(file.filename):
        # Create secure filename with claim_id prefix
        filename = secure_filename(file.filename)
        name, ext = os.path.splitext(filename)
        secure_filename_with_id = f"{claim_id}_{name}{ext}"
        
        # Ensure upload folder exists
        os.makedirs(upload_folder, exist_ok=True)
        
        # Save file
        file_path = os.path.join(upload_folder, secure_filename_with_id)
        file.save(file_path)
        return file_path
    return None


def parse_month_year(month_year_str):
    """Parse month-year string like '2024-01' into month and year integers"""
    try:
        if not month_year_str or '-' not in month_year_str:
            return None, None
        
        year_str, month_str = month_year_str.split('-')
        month = int(month_str)
        year = int(year_str)
        
        # Validate month range
        if 1 <= month <= 12 and year > 0:
            return month, year
        else:
            return None, None
    except (ValueError, AttributeError):
        return None, None


def get_available_months():
    """Get list of available months that have claims"""
    from models import Claim
    
    # Get all claims with their from_date
    claims = Claim.query.all()
    
    if not claims:
        return []
    
    # Extract unique year-month combinations
    months = set()
    for claim in claims:
        if claim.from_date:
            year_month = claim.from_date.strftime('%Y-%m')
            months.add(year_month)
    
    # Sort months and create choices list
    sorted_months = sorted(months, reverse=True)  # Most recent first
    
    # Format for display
    choices = []
    for month in sorted_months:
        year, month_num = month.split('-')
        month_names = [
            '', 'January', 'February', 'March', 'April', 'May', 'June',
            'July', 'August', 'September', 'October', 'November', 'December'
        ]
        display_name = f"{month_names[int(month_num)]} {year}"
        choices.append((month, display_name))
    
    return choices


def get_available_claims():
    """Get list of available claims for selection"""
    from models import Claim
    
    claims = Claim.query.order_by(Claim.from_date.desc()).all()
    claim_choices = []
    
    for claim in claims:
        # Create a descriptive label for each claim
        date_range = f"{claim.from_date.strftime('%m/%d/%Y')}"
        if claim.from_date != claim.to_date:
            date_range += f" - {claim.to_date.strftime('%m/%d/%Y')}"
        
        alias_part = f" ({claim.alias_name})" if claim.alias_name else ""
        amount_part = f" - {claim.total_currency} {claim.total_amount}"
        group_part = f" [{claim.expense_group}]"
        
        label = f"{date_range}{alias_part}{amount_part}{group_part}"
        value = claim.claim_id
        
        claim_choices.append((value, label))
    
    return claim_choices


def get_months_from_claims(claim_ids):
    """Get unique months from selected claims"""
    from models import Claim
    
    claims = Claim.query.filter(Claim.claim_id.in_(claim_ids)).all()
    months = set()
    
    for claim in claims:
        year = claim.from_date.year
        month = claim.from_date.month
        months.add((year, month))
    
    return sorted(months)


def generate_isd_reimbursement_csv(month_year_str):
    """Generate CSV for ISD Reimbursement Form (individual items)"""
    from models import Claim, ClaimItem
    
    month, year = parse_month_year(month_year_str)
    if not month or not year:
        raise ValueError("Invalid month/year format")
    
    # Query items for the specified month/year
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)
    
    items = ClaimItem.query.join(Claim).filter(
        Claim.from_date >= start_date.date(),
        Claim.from_date < end_date.date()
    ).order_by(ClaimItem.created_at).all()
    
    # Create CSV content
    output = StringIO()
    writer = csv.writer(output)
    
    # Find all currencies used in this month to determine the "Others" column
    currencies_used = set(item.currency for item in items)
    other_currencies = currencies_used - {'HKD', 'RMB'}
    other_currency = list(other_currencies)[0] if len(other_currencies) == 1 else 'Others'
    
    # Write header
    writer.writerow([
        'Receipt Order',
        'Payment Date',
        'Particulars',
        'HKD ($)',
        'RMB ($)',
        f'{other_currency} ($)' if other_currency != 'Others' else 'Others ($)',
        'Expense Group',
        'Receipt Attached?'
    ])
    
    # Initialize totals
    hkd_total = 0
    rmb_total = 0
    other_total = 0
    
    # Write data rows
    for i, item in enumerate(items, 1):
        payment_date = item.claim.from_date.strftime('%d-%m-%Y')
        
        # Determine amounts by currency
        hkd_amount = ''
        rmb_amount = ''
        other_amount = ''
        
        if item.currency == 'HKD':
            hkd_amount = str(item.amount)
            hkd_total += float(item.amount)
        elif item.currency == 'RMB':
            rmb_amount = str(item.amount)
            rmb_total += float(item.amount)
        else:
            other_amount = str(item.amount)
            other_total += float(item.amount)
        
        receipt_attached = 'Yes' if item.claim.upload_file_path else 'No'
        expense_group = item.claim.expense_group
        
        writer.writerow([
            i,
            payment_date,
            item.description,
            hkd_amount,
            rmb_amount,
            other_amount,
            expense_group,
            receipt_attached
        ])
    
    # Add totals row
    writer.writerow([
        '',
        '',
        'TOTAL',
        f'{hkd_total:.2f}' if hkd_total > 0 else '',
        f'{rmb_total:.2f}' if rmb_total > 0 else '',
        f'{other_total:.2f}' if other_total > 0 else '',
        '',
        ''
    ])
    
    return output.getvalue()


def generate_multi_claim_isd_reports(claim_ids):
    """Generate ISD reports for multiple claims, one per month"""
    from models import Claim, ClaimItem
    
    # Get unique months from selected claims
    months = get_months_from_claims(claim_ids)
    reports = {}
    
    for year, month in months:
        # Query items for this specific month from selected claims
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
        
        items = ClaimItem.query.join(Claim).filter(
            Claim.claim_id.in_(claim_ids),
            Claim.from_date >= start_date.date(),
            Claim.from_date < end_date.date()
        ).order_by(ClaimItem.created_at).all()
        
        if items:  # Only create report if there are items for this month
            # Create CSV content
            output = StringIO()
            writer = csv.writer(output)
            
            # Find all currencies used in this month to determine the "Others" column
            currencies_used = set(item.currency for item in items)
            other_currencies = currencies_used - {'HKD', 'RMB'}
            other_currency = list(other_currencies)[0] if len(other_currencies) == 1 else 'Others'
            
            # Write header
            writer.writerow([
                'Receipt Order',
                'Payment Date',
                'Particulars',
                'HKD ($)',
                'RMB ($)',
                f'{other_currency} ($)' if other_currency != 'Others' else 'Others ($)',
                'Expense Group',
                'Receipt Attached?'
            ])
            
            # Initialize totals
            hkd_total = 0
            rmb_total = 0
            other_total = 0
            
            # Write data rows
            for i, item in enumerate(items, 1):
                payment_date = item.claim.from_date.strftime('%d-%m-%Y')
                
                # Determine amounts by currency
                hkd_amount = ''
                rmb_amount = ''
                other_amount = ''
                
                if item.currency == 'HKD':
                    hkd_amount = str(item.amount)
                    hkd_total += float(item.amount)
                elif item.currency == 'RMB':
                    rmb_amount = str(item.amount)
                    rmb_total += float(item.amount)
                else:
                    other_amount = str(item.amount)
                    other_total += float(item.amount)
                
                receipt_attached = 'Yes' if item.claim.upload_file_path else 'No'
                expense_group = item.claim.expense_group
                
                writer.writerow([
                    i,
                    payment_date,
                    item.description,
                    hkd_amount,
                    rmb_amount,
                    other_amount,
                    expense_group,
                    receipt_attached
                ])
            
            # Add totals row
            writer.writerow([
                '',
                '',
                'TOTAL',
                f'{hkd_total:.2f}' if hkd_total > 0 else '',
                f'{rmb_total:.2f}' if rmb_total > 0 else '',
                f'{other_total:.2f}' if other_total > 0 else '',
                '',
                ''
            ])
            
            month_key = f"{year:04d}_{month:02d}"
            reports[month_key] = {
                'filename': f'isd_reimbursement_{year}_{month:02d}.csv',
                'content': output.getvalue(),
                'month_name': datetime(year, month, 1).strftime('%B %Y')
            }
    
    return reports


def generate_financial_expense_csv(month_year_str=None):
    """Generate CSV for Financial Office Expense Form (claims)"""
    from models import Claim
    
    # Query claims for the specified month/year or all if not specified
    query = Claim.query
    
    if month_year_str:
        month, year = parse_month_year(month_year_str)
        if month and year:
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year + 1, 1, 1)
            else:
                end_date = datetime(year, month + 1, 1)
            
            query = query.filter(
                Claim.from_date >= start_date.date(),
                Claim.from_date < end_date.date()
            )
    
    claims = query.order_by(Claim.created_at).all()
    
    # Create CSV content
    output = StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'Incurred Date From',
        'Incurred Date To',
        'Description',
        'Paid Currency',
        'Paid Total Amount',
        'Expense Group',
        'Alias Name',
        'Business Purpose',
        'Justifications',
        'UUID'
    ])
    
    # Write data rows
    for claim in claims:
        # Combine all item descriptions
        descriptions = '; '.join([item.description for item in claim.items])
        
        # Combine all justifications
        justifications = '; '.join([item.justification for item in claim.items if item.justification])
        
        paid_currency = claim.paid_currency or claim.total_currency
        paid_amount = claim.paid_amount or claim.total_amount
        alias_name = claim.alias_name or ''
        
        writer.writerow([
            claim.from_date.strftime('%d-%m-%Y'),
            claim.to_date.strftime('%d-%m-%Y'),
            descriptions,
            paid_currency,
            str(paid_amount),
            claim.expense_group,
            alias_name,
            claim.business_purpose,
            justifications,
            claim.claim_id
        ])
    
    return output.getvalue()


def generate_multi_claim_financial_csv(claim_ids):
    """Generate combined Financial Office Expense CSV for multiple claims"""
    from models import Claim
    
    # Query selected claims
    claims = Claim.query.filter(Claim.claim_id.in_(claim_ids)).order_by(Claim.created_at).all()
    
    # Create CSV content
    output = StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'Incurred Date From',
        'Incurred Date To',
        'Description',
        'Paid Currency',
        'Paid Total Amount',
        'Expense Group',
        'Alias Name',
        'Business Purpose',
        'Justifications',
        'UUID'
    ])
    
    # Write data rows
    for claim in claims:
        # Combine all item descriptions
        descriptions = '; '.join([item.description for item in claim.items])
        
        # Combine all justifications
        justifications = '; '.join([item.justification for item in claim.items if item.justification])
        
        paid_currency = claim.paid_currency or claim.total_currency
        paid_amount = claim.paid_amount or claim.total_amount
        alias_name = claim.alias_name or ''
        
        writer.writerow([
            claim.from_date.strftime('%d-%m-%Y'),
            claim.to_date.strftime('%d-%m-%Y'),
            descriptions,
            paid_currency,
            str(paid_amount),
            claim.expense_group,
            alias_name,
            claim.business_purpose,
            justifications,
            claim.claim_id
        ])
    
    return output.getvalue()


def create_receipts_zip(month_year_str=None, upload_folder='uploads'):
    """Create ZIP file with all receipts for specified period"""
    from models import Claim
    
    # Query claims for the specified month/year or all if not specified
    query = Claim.query
    
    if month_year_str:
        month, year = parse_month_year(month_year_str)
        if month and year:
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year + 1, 1, 1)
            else:
                end_date = datetime(year, month + 1, 1)
            
            query = query.filter(
                Claim.from_date >= start_date.date(),
                Claim.from_date < end_date.date()
            )
    
    claims = query.all()
    
    # Create ZIP file
    zip_path = f"receipts_{month_year_str.replace('-', '_')}.zip" if month_year_str else "receipts_all.zip"
    
    with zipfile.ZipFile(zip_path, 'w') as zip_file:
        for claim in claims:
            if claim.upload_file_path and os.path.exists(claim.upload_file_path):
                # Get original filename and add claim ID
                original_filename = os.path.basename(claim.upload_file_path)
                name, ext = os.path.splitext(original_filename)
                alias_prefix = f"{claim.alias_name}_" if claim.alias_name else ""
                zip_filename = f"receipt_{alias_prefix}{claim.claim_id}{ext}"
                
                zip_file.write(claim.upload_file_path, zip_filename)
    
    return zip_path


def create_multi_claim_receipts_zip(claim_ids, upload_folder='uploads'):
    """Create ZIP file with receipts from selected claims"""
    from models import Claim
    
    # Query selected claims
    claims = Claim.query.filter(Claim.claim_id.in_(claim_ids)).all()
    
    # Create timestamp for unique filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    zip_path = f"receipts_multi_claim_{timestamp}.zip"
    
    with zipfile.ZipFile(zip_path, 'w') as zip_file:
        attachment_counter = 1
        
        for claim in claims:
            if claim.upload_file_path and os.path.exists(claim.upload_file_path):
                # Get original file extension
                original_filename = os.path.basename(claim.upload_file_path)
                name, ext = os.path.splitext(original_filename)
                
                # Create new filename: uuid_Attachment01.pdf format
                zip_filename = f"{claim.claim_id}_Attachment{attachment_counter:02d}{ext}"
                
                zip_file.write(claim.upload_file_path, zip_filename)
                attachment_counter += 1
    
    return zip_path


def create_multi_report_zip(claim_ids):
    """Create a comprehensive ZIP with ISD reports per month, combined financial report, and receipts"""
    import tempfile
    import shutil
    
    # Create timestamp for unique filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    zip_path = f"comprehensive_report_{timestamp}.zip"
    
    with zipfile.ZipFile(zip_path, 'w') as zip_file:
        # 1. Generate ISD reports per month
        isd_reports = generate_multi_claim_isd_reports(claim_ids)
        for month_key, report_data in isd_reports.items():
            zip_file.writestr(report_data['filename'], report_data['content'])
        
        # 2. Generate combined financial report
        financial_csv = generate_multi_claim_financial_csv(claim_ids)
        zip_file.writestr('financial_expense_combined.csv', financial_csv)
        
        # 3. Add receipts with new naming convention
        from models import Claim
        claims = Claim.query.filter(Claim.claim_id.in_(claim_ids)).all()
        
        attachment_counter = 1
        for claim in claims:
            if claim.upload_file_path and os.path.exists(claim.upload_file_path):
                # Get original file extension
                original_filename = os.path.basename(claim.upload_file_path)
                name, ext = os.path.splitext(original_filename)
                
                # Create new filename: uuid_Attachment01.pdf format
                zip_filename = f"receipts/{claim.claim_id}_Attachment{attachment_counter:02d}{ext}"
                
                zip_file.write(claim.upload_file_path, zip_filename)
                attachment_counter += 1
    
    return zip_path


def generate_excel_isd_report(claims_or_month, template_path='reference/isd_template.xlsx'):
    """Generate ISD reimbursement report directly in Excel format using template.
    
    Args:
        claims_or_month: Either a list of Claim objects or a month string (e.g., '2025-05')
        template_path: Path to the Excel template file
    """
    try:
        import openpyxl
        from openpyxl import load_workbook
        from openpyxl.styles import Border, Side, Alignment, Font
        from openpyxl.utils import get_column_letter
        from decimal import Decimal
        import io
        
        # Handle both claim objects and month string input
        if isinstance(claims_or_month, str):
            # It's a month string, need to query for claims
            from models import Claim, db
            month, year = parse_month_year(claims_or_month)
            if not month or not year:
                return None
            
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year + 1, 1, 1)
            else:
                end_date = datetime(year, month + 1, 1)
            
            claims = Claim.query.filter(
                Claim.from_date >= start_date.date(),
                Claim.from_date < end_date.date()
            ).all()
        else:
            # It's already a list of claims
            claims = claims_or_month
        
        # Load the template
        wb = load_workbook(template_path)
        ws = wb.active
        
        if not claims:
            return None
        
        # Get the first claim for header information
        first_claim = claims[0]
        
        # Populate header information
        ws['C3'] = 'SE'  # SE Number - could be made configurable
        ws['C4'] = getattr(first_claim, 'claimant_name', '')
        ws['C5'] = getattr(first_claim, 'claimant_id', '')
        ws['C6'] = getattr(first_claim, 'claimant_email', '')
        
        # Collect all items from all claims
        all_items = []
        for claim in claims:
            for item in claim.items:
                all_items.append({
                    'claim': claim,
                    'item': item
                })
        
        # Sort items by date
        all_items.sort(key=lambda x: x['claim'].from_date)
        
        # Group items by month for the template structure
        from collections import defaultdict
        items_by_month = defaultdict(list)
        for item_data in all_items:
            month_key = item_data['claim'].from_date.strftime('%Y-%m')
            items_by_month[month_key].append(item_data)
        
        # Template structure analysis:
        # Month 1: Period at row 10, Headers at row 11, Data rows 12-13, Total at row 14
        # Month 2: Period at row 16, Headers at row 17, Data rows 18-19, Total at row 20
        
        # Clear existing template data first, but preserve structure
        template_sections = [
            {'period_row': 10, 'header_row': 11, 'data_start': 12, 'data_end': 13, 'total_row': 14},
            {'period_row': 16, 'header_row': 17, 'data_start': 18, 'data_end': 19, 'total_row': 20}
        ]
        
        # Clear template example data
        for section in template_sections:
            for row in range(section['data_start'], section['data_end'] + 1):
                for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    ws[f'{col}{row}'].value = None
        
        # Store original formatting from template data row
        template_data_row = 12
        original_styles = {}
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            cell = ws[f'{col}{template_data_row}']
            original_styles[col] = {
                'font': Font(name=cell.font.name or 'Calibri', size=cell.font.size or 11, bold=cell.font.bold),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                ),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'number_format': cell.number_format
            }
        
        # Special alignment for particulars column (left-aligned)
        original_styles['D']['alignment'] = Alignment(horizontal='left', vertical='center')
        
        # Now populate data section by section
        sorted_months = sorted(items_by_month.keys())
        
        current_section_idx = 0
        
        for month_key in sorted_months:
            month_items = items_by_month[month_key]
            month_date = datetime.strptime(month_key + '-01', '%Y-%m-%d')
            
            if current_section_idx < len(template_sections):
                # Use existing template section
                section = template_sections[current_section_idx]
                
                # Update period
                ws[f'C{section["period_row"]}'] = month_date.strftime('%B %Y')
                
                # Calculate how many rows we need for this month's data
                rows_needed = len(month_items)
                available_rows = section['data_end'] - section['data_start'] + 1
                
                if rows_needed > available_rows:
                    # Need to insert more rows in this section
                    rows_to_insert = rows_needed - available_rows
                    ws.insert_rows(section['data_end'] + 1, rows_to_insert)
                    
                    # Update ALL subsequent sections' row numbers (critical fix!)
                    for i in range(current_section_idx + 1, len(template_sections)):
                        for key in template_sections[i]:
                            template_sections[i][key] += rows_to_insert
                    
                    # Update current section's data_end
                    section['data_end'] += rows_to_insert
                    # Update total row
                    section['total_row'] += rows_to_insert
                
                # Populate data for this month
                data_row = section['data_start']
                
            else:
                # Need to create a new section at the end
                # Find where to place the new section (after the last used section)
                if template_sections:  # Check if we have any sections at all
                    # Use the last defined section as reference
                    last_section_idx = min(current_section_idx - 1, len(template_sections) - 1)
                    if last_section_idx >= 0:
                        prev_section = template_sections[last_section_idx]
                        insertion_point = prev_section['total_row'] + 2  # After total + 1 spacing row
                    else:
                        insertion_point = 26  # Default end position
                else:
                    insertion_point = 26  # Default end position
                
                # Insert rows for: 1 spacing + 1 period + 1 header + data rows + 1 total + 1 spacing
                total_rows_needed = 1 + 1 + 1 + len(month_items) + 1 + 1
                ws.insert_rows(insertion_point, total_rows_needed)
                
                section = {
                    'period_row': insertion_point + 1,
                    'header_row': insertion_point + 2,
                    'data_start': insertion_point + 3,
                    'data_end': insertion_point + 2 + len(month_items),
                    'total_row': insertion_point + 3 + len(month_items)
                }
                
                # Add this new section to our template_sections list for future reference
                template_sections.append(section)
                
                # Set up the new section
                ws[f'B{section["period_row"]}'] = 'Period:'
                ws[f'C{section["period_row"]}'] = month_date.strftime('%B %Y')
                
                # Copy headers from existing section (use the first section as template)
                if template_sections:
                    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                        template_header_row = 11  # Use the original header row as template
                        ws[f'{col}{section["header_row"]}'] = ws[f'{col}{template_header_row}'].value
                
                data_row = section['data_start']
            
            # Populate items for this month
            for idx, item_data in enumerate(month_items, 1):
                item = item_data['item']
                claim = item_data['claim']
                
                # Apply formatting to all cells in this row
                for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    target_cell = ws[f'{col}{data_row}']
                    style = original_styles[col]
                    target_cell.font = style['font']
                    target_cell.border = style['border']
                    target_cell.alignment = style['alignment']
                    target_cell.number_format = style['number_format']
                
                # Receipt Order
                ws[f'B{data_row}'] = idx
                
                # Payment Date
                ws[f'C{data_row}'] = claim.from_date.strftime('%d-%m-%Y')
                
                # Particulars
                ws[f'D{data_row}'] = item.description
                
                # Currency amounts - place in appropriate column
                if item.currency == 'HKD':
                    ws[f'E{data_row}'] = float(item.amount)
                elif item.currency == 'RMB':
                    ws[f'F{data_row}'] = float(item.amount)
                else:
                    # For other currencies (EUR, USD, GBP, JPY), use Others column
                    ws[f'G{data_row}'] = float(item.amount)
                    # Update the header to show the specific currency
                    ws[f'G{section["header_row"]}'] = f'Others (Specify:{item.currency})'
                
                # Receipt Attached
                ws[f'H{data_row}'] = 'Yes' if claim.upload_file_path else 'No'
                
                data_row += 1
            
            # Add totals row for this month
            if month_items:
                total_row = section['total_row']
                
                # Apply formatting to totals row
                for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    target_cell = ws[f'{col}{total_row}']
                    style = original_styles[col]
                    target_cell.font = Font(bold=True, name=style['font'].name, size=style['font'].size)
                    target_cell.border = style['border']
                    target_cell.alignment = style['alignment']
                    target_cell.number_format = style['number_format']
                
                # Clear any existing formulas/values
                for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    ws[f'{col}{total_row}'].value = None
                
                # TOTAL row content
                ws[f'D{total_row}'] = 'Total:'
                
                # Calculate totals for this month
                hkd_total = sum(float(item_data['item'].amount) for item_data in month_items 
                               if item_data['item'].currency == 'HKD')
                if hkd_total > 0:
                    ws[f'E{total_row}'] = hkd_total
                
                rmb_total = sum(float(item_data['item'].amount) for item_data in month_items 
                               if item_data['item'].currency == 'RMB')
                if rmb_total > 0:
                    ws[f'F{total_row}'] = rmb_total
                
                others_total = sum(float(item_data['item'].amount) for item_data in month_items 
                                  if item_data['item'].currency not in ['HKD', 'RMB'])
                if others_total > 0:
                    ws[f'G{total_row}'] = others_total
            
            current_section_idx += 1
        
        # Clean up any unused template sections
        while current_section_idx < len(template_sections):
            section = template_sections[current_section_idx]
            
            # Clear the unused section completely
            for row in range(section['period_row'], section['total_row'] + 2):  # +2 for spacing
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                    cell = ws[f'{col}{row}']
                    if cell.value:
                        cell.value = None
            
            current_section_idx += 1
        
        # Auto-adjust column widths for better visibility
        column_widths = {
            'B': 12,  # Receipt Order
            'C': 15,  # Payment Date
            'D': 30,  # Particulars
            'E': 12,  # HKD
            'F': 12,  # RMB
            'G': 15,  # Others
            'H': 18   # Receipt Attached
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Save to BytesIO buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
        
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        return None
    except Exception as e:
        print(f"Error generating Excel report: {e}")
        return None


def generate_multi_claim_excel_reports(claim_ids):
    """Generate Excel ISD reports for multiple claims, grouped by month."""
    from models import Claim
    
    # Get claims
    claims = Claim.query.filter(Claim.claim_id.in_(claim_ids)).all()
    if not claims:
        return {}
    
    # Group claims by month
    monthly_claims = {}
    for claim in claims:
        month_key = claim.from_date.strftime('%Y-%m')
        if month_key not in monthly_claims:
            monthly_claims[month_key] = []
        monthly_claims[month_key].append(claim)
    
    # Generate Excel report for each month
    reports = {}
    for month_key, month_claims in monthly_claims.items():
        month_date = datetime.strptime(month_key, '%Y-%m')
        month_name = month_date.strftime('%B %Y')
        
        excel_content = generate_excel_isd_report(month_claims)
        if excel_content:
            reports[month_key] = {
                'month_name': month_name,
                'content': excel_content,
                'filename': f'ISD_Reimbursement_{month_name.replace(" ", "_")}.xlsx'
            }
    
    return reports


def create_multi_report_excel_zip(claim_ids):
    """Create a comprehensive ZIP with single Excel ISD report (all months), combined financial report, and receipts"""
    import tempfile
    
    # Create timestamp for unique filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    zip_path = f"comprehensive_excel_report_{timestamp}.zip"
    
    with zipfile.ZipFile(zip_path, 'w') as zip_file:
        # 1. Generate single comprehensive Excel ISD report with all months
        from models import Claim
        claims = Claim.query.filter(Claim.claim_id.in_(claim_ids)).all()
        
        if claims:
            # Generate single Excel file with all claims across all months
            excel_content = generate_excel_isd_report(claims)
            if excel_content:
                # Create descriptive filename based on date range
                dates = [claim.from_date for claim in claims]
                start_date = min(dates).strftime('%Y-%m')
                end_date = max(dates).strftime('%Y-%m')
                
                if start_date == end_date:
                    filename = f'ISD_Comprehensive_Report_{start_date}.xlsx'
                else:
                    filename = f'ISD_Comprehensive_Report_{start_date}_to_{end_date}.xlsx'
                
                zip_file.writestr(filename, excel_content)
        
        # 2. Generate combined financial report (still CSV for compatibility)
        financial_csv = generate_multi_claim_financial_csv(claim_ids)
        zip_file.writestr('financial_expense_combined.csv', financial_csv)
        
        # 3. Add receipts with new naming convention
        attachment_counter = 1
        for claim in claims:
            if claim.upload_file_path and os.path.exists(claim.upload_file_path):
                # Get original file extension
                original_filename = os.path.basename(claim.upload_file_path)
                name, ext = os.path.splitext(original_filename)
                
                # Create new filename: uuid_Attachment01.pdf format
                zip_filename = f"receipts/{claim.claim_id}_Attachment{attachment_counter:02d}{ext}"
                
                zip_file.write(claim.upload_file_path, zip_filename)
                attachment_counter += 1
    
    return zip_path